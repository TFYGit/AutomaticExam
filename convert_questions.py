"""
将题库 Excel 文件转换为 questions.json
支持格式：
  - Sheet 1/3/4: [题目类型, 题目, 正确答案, 答案A, 答案B, 答案C, 答案D, 答案E, 答案F]
  - Sheet 2:     [序号, 题目类型, 题目, meta, meta, meta, 正确答案, 答案A-F]
"""
import re
import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r'C:\Users\GG\Desktop\题库-2026年3月更新版(1).xlsx'
OUTPUT_PATH = r'D:\pycharm\exam-system\src\data\questions.json'

VALID_TYPES = {'单选题', '多选题', '判断题', '判断图'}
TYPE_MAP = {
    '单选题': 'choice',
    '多选题': 'multichoice',
    '判断题': 'truefalse',
    '判断图': 'truefalse',
}
OPT_KEYS = ['A', 'B', 'C', 'D', 'E', 'F']


def is_cjk(c):
    # CJK 统一汉字、扩展A、CJK 符号和标点（含顿号、句号等）
    return ('\u4e00' <= c <= '\u9fff' or '\u3400' <= c <= '\u4dbf'
            or '\u3000' <= c <= '\u303f' or c in '，。！？；：""''（）【】')

def clean_text(val):
    if val is None:
        return None
    text = str(val).strip()
    # 中文字符之间的换行直接删掉（不加空格），其他位置换行替换为空格
    result = []
    i = 0
    while i < len(text):
        ch = text[i]
        if ch in ('\n', '\r'):
            prev = result[-1] if result else ''
            next_ch = text[i + 1] if i + 1 < len(text) else ''
            if is_cjk(prev) and is_cjk(next_ch):
                pass  # 直接跳过换行，中文间无需空格
            else:
                result.append(' ')
        else:
            result.append(ch)
        i += 1
    text = ''.join(result).strip()
    # 合并多余空格
    text = re.sub(r' {2,}', ' ', text)
    return text if text else None


def parse_answer(raw, qtype):
    if raw is None:
        return None
    s = str(raw).strip()
    if qtype == 'truefalse':
        return '对' if s in ('正确', '对', 'True', 'true', '是') else '错'
    if qtype == 'multichoice':
        letters = [c.upper() for c in s if c.isalpha()]
        return letters if letters else None
    # choice
    return s.upper() if s else None


def detect_cols(ws):
    """根据第二行第一个单元格判断列结构，返回 (type_col, question_col, answer_col, opt_start_col)"""
    cell1 = ws.cell(2, 1).value
    if isinstance(cell1, (int, float)) and cell1 is not None:
        # Sheet 2 格式：序号在首列
        return 2, 3, 7, 8
    else:
        # 普通格式
        return 1, 2, 3, 4


def convert():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    questions = []
    q_id = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        type_col, question_col, answer_col, opt_start_col = detect_cols(ws)
        print(f'Sheet: {sheet_name}  rows={ws.max_row}  '
              f'cols=({type_col},{question_col},{answer_col},{opt_start_col})')

        for row in range(2, ws.max_row + 1):
            qtype_raw = ws.cell(row, type_col).value
            if qtype_raw not in VALID_TYPES:
                continue

            qtype = TYPE_MAP[qtype_raw]
            question = clean_text(ws.cell(row, question_col).value)
            answer_raw = ws.cell(row, answer_col).value

            if not question:
                continue

            answer = parse_answer(answer_raw, qtype)
            if answer is None:
                continue

            # 构建选项
            options = {}
            for i, key in enumerate(OPT_KEYS):
                col = opt_start_col + i
                if col > ws.max_column:
                    break
                val = ws.cell(row, col).value
                if val is not None:
                    text = clean_text(val)
                    if text:
                        options[key] = text

            q = {
                'id': q_id,
                'type': qtype,
                'question': question,
                'answer': answer,
                'explanation': '',
            }
            if qtype != 'truefalse':
                q['options'] = options

            questions.append(q)
            q_id += 1

    print(f'\n共转换 {len(questions)} 道题')
    for t in ['choice', 'multichoice', 'truefalse']:
        cnt = sum(1 for q in questions if q['type'] == t)
        label = {'choice': '单选题', 'multichoice': '多选题', 'truefalse': '判断题'}[t]
        print(f'  {label}: {cnt}')

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(questions, f, ensure_ascii=False, indent=2)
    print(f'\n已保存至 {OUTPUT_PATH}')


if __name__ == '__main__':
    convert()
